import shutil
from datetime import timedelta, datetime
import time
import hashlib
import math
import socket
import openpyxl
import json
import os
import io
from requests import get, adapters
from urllib3 import poolmanager
from ssl import create_default_context, Purpose, CERT_NONE
from requests import Session
from openpyxl.utils import get_column_letter
import zipfile
from pyhtml2pdf import converter
# import uuid, wmi
from openpyxl.styles import Border, Side, Alignment, Font
from pymongo import MongoClient
from datetime import datetime

class InvoiceLogic:
    def __init__(self):
        self.download_path = '.\\__pycache__\\cache_'
        self.full_path = os.path.abspath(self.download_path)
        self.pb_t = "v3.6.4 (03/03/2025)"
        self.ch = 0
        self.true_acc = True
        self.token_ = 0
        self.headers = None
        self.user = None
        self.pass_ = None
        self.ckey = None
        self.captcha_inp = None
        self.checkspc = " "
        self.arr_ed = None
        self.begin_ = None
        self.end_ = None
        self.range_DAY_const = None
        self.path_ketqua = None
        self.banra = False
        self.muavao = False
        self.tongquat = False
        self.chitiet = False
        self.xml_ = False
        self.html_ = False
        self.xmltpdf = False
        # MongoDB connection - Atlas only
        self.mongo_client = MongoClient("mongodb+srv://AdminMola:MolaMola123@molanpl.gplydvn.mongodb.net/")
        self.db = self.mongo_client["MolaDatabase"]
        self.sales_collection = self.db["HoaDonBanRa"]
        self.purchase_collection = self.db["HoaDonMuaVao"]

    def ssl_supressed_session(self):
        ctx = create_default_context(Purpose.SERVER_AUTH)
        ctx.check_hostname = False
        ctx.verify_mode = CERT_NONE
        ctx.options |= 0x4    
        session = Session()
        session.mount('https://', CustomHttpAdapter(ctx))
        return session

    def get_current_date(self):
        ngay_hien_tai = datetime.now()
        ngay_thang_nam = ngay_hien_tai.strftime('%d/%m/%Y')
        self.ngtn = ngay_thang_nam
        return ngay_thang_nam

    def compare_dates(self, date1, date2):
        format = "%d/%m/%Y"
        datetime1 = datetime.strptime(date1, format)
        datetime2 = datetime.strptime(date2, format)
        if datetime1 < datetime2:
            print(0)
            return 0
        return 1

    def increase_date(self, date_string):
        try:
            date = datetime.strptime(date_string, "%d/%m/%Y")
            increased_date = date + timedelta(days=1)
            increased_date_string = increased_date.strftime("%d/%m/%Y")
            return increased_date_string
        except ValueError:
            return "Định dạng ngày không hợp lệ!"

    def day_dow(self, start_date, end_date):
        print("=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=-=")
        print(start_date)
        print(end_date)
        date_format = "%d/%m/%Y"
        try:
            date1 = datetime.strptime(start_date, date_format)
            date2 = datetime.strptime(end_date, date_format)
            
            # Kiểm tra nếu ngày bắt đầu lớn hơn ngày kết thúc
            if date1 > date2:
                print("Lỗi: Ngày bắt đầu lớn hơn ngày kết thúc")
                return [['01/01/2024', '31/01/2024']]  # Trả về mảng mặc định
                
            one_month = timedelta(days=27)
            date_ranges = []
            while date1 <= date2:
                sub_array = []
                sub_array.append(date1.strftime(date_format))
                date1 += one_month
                if date1 > date2:
                    date1 = date2
                sub_array.append(date1.strftime(date_format))
                date_ranges.append(sub_array)
                date1 += timedelta(days=1)
            
            # Nếu mảng rỗng, trả về mảng mặc định
            if not date_ranges:
                print("Không thể tạo khoảng thời gian, sử dụng giá trị mặc định")
                return [['01/01/2024', '31/01/2024']]
                
            print(date_ranges)
            return date_ranges
        except ValueError as e:
            print(f"Lỗi định dạng ngày: {e}")
            # Trả về mảng mặc định
            return [['01/01/2024', '31/01/2024']]
        except Exception as e:
            print(f"Lỗi trong quá trình xử lý ngày: {e}")
            # Trả về mảng mặc định
            return [['01/01/2024', '31/01/2024']]

    def remove_duplicate_elements(self, data):
        seen_elements = set()
        unique_elements = []
        for element in data:
            element_json = json.dumps(element, sort_keys=True)
            if element_json not in seen_elements:
                seen_elements.add(element_json)
                unique_elements.append(element)
        unique_json_array = unique_elements
        return unique_json_array

    def login_web(self):
        payload = { 
            'ckey': self.ckey,
            'cvalue': self.captcha_inp,
            'password': self.pass_,
            'username': self.user
        } 
        login_to = self.ssl_supressed_session().post('https://hoadondientu.gdt.gov.vn:30000/security-taxpayer/authenticate',verify=False,json=payload).json()
        print(login_to)
        self.token_ = 0
        self.true_acc = True
        if 'token' in login_to and self.true_acc:
            print('Login success')
            self.token_ = login_to['token']
            self.headers = {
                "Authorization":'Bearer ' + self.token_ 
            }       
            return True
        return False

    def check_user(self):
        with open(".\\__pycache__\\cache_\\user.txt", "r") as file:
            line = file.readline()
            line = line.strip()
            self.user, self.password = line.strip().split("|")
        return self.user, self.password

    def save_user(self):
        with open(".\\__pycache__\\cache_\\user.txt", "w") as file:
            file.write(f"{self.user.rstrip()}|{self.pass_.rstrip()}")

    def chitiet_(self, type):
        list_link = [['0100684378', 'https://0100684378-tt78.vnpt-invoice.com.vn'],['0314743623', 'https://ehoadondientu.com/Tra-cuu\n'],['0105987432', 'EASy'],['0106741551', 'https://tracuuhoadon.cyberbill.vn/#/tracuuhoadon/tracuu'],['0102721191-068', 'https://vat.ggg.com.vn/'],['0102516308', 'https://tracuuhoadon.mediamart.com.vn/'],['0107500414', 'https://tracuuhoadon.vetc.com.vn/'],['4600128263', 'https://hoadon.petrolimex.com.vn/SearchInvoicebycode/Index'],['0100107564-001', 'https://hoadon.petrolimex.com.vn/SearchInvoicebycode/Index'], ['0104128565', 'https://bit.ly/hdtrcuuFPT\n'], ['0302999571', 'https://tracuu.lcs-ca.vn\n'], ['0313963672', 'https://tracuuhoadon.kkvat.com.vn/\n'], ['0105232093', 'https://tracuu.cyberbill.vn/#/tracuuhoadon/tracuu\n'], ['0311942758', 'http://www.ngogiaphat.vn ( website tra cứu không thể truy cập )\n'], ['0302712571', 'https://matbao.in/tra-cuu-hoa-don/\n'], ['0103930279', 'https://www.nacencomm.com.vn/dich-vu-chi-tiet/hoa-don-dien-tu\n'], ['0105844836', 'https://tracuu.vininvoice.vn\n'], ['0312483391', 'https://azinvoice.com ( web tra cứu lỗi )\n'], ['0101243150', 'https://www.meinvoice.vn/tra-cuu/\n'], ['0106026495', 'https://tracuuhoadon.minvoice.com.vn/single/invoice\n'], ['0313906508', 'www.nguyenminhvat.vn\n'], ['0101300842', 'https://einvoice.vn/tra-cuu\n'], ['0306784030', 'https://ehoadon.online/einvoice/lookup\n'], ['0200638946', 'https://oinvoice.vn/tracuu/\n'], ['0312303803', 'https://tracuu.wininvoice.vn\n'], ['0100109106', 'https://vinvoice.viettel.vn/utilities/invoice-search\n'], ['0102454468', 'https://tax24.com.vn/thuedientu/xac-minh-hoa-don\n'], ['0105937449', 'https://newinvoice.com.vn/tra-cuu/\n'], ['0108516079', 'http://hddt.3asoft.vn/#tracuu\n'], ['0100686209', 'https://bit.ly/hdtrcuumobifone\n'], ['0101360697', 'https://bit.ly/hdtracuuVan\n'], ['0101162173', 'https://asiainvoice.vn/tra-cuu\n'], ['0401486901', 'https://tracuu.vin-hoadon.com/tracuuhoadon/tracuuxacthuc/tracuuhd\n'], ['0200784873', 'https://dinhvibachkhoa.vn\n'], ['0100684378', 'https://portaltool-miennam.vnpt-invoice.com.vn\n'], ['0106713804', 'https://hiloinvoice.vn/tra-cuu/\n'], ['0314209362', 'https://hoadondientuvat.com/Tracuu.aspx\n'], ['0101352495', 'https://tracuu.v50.vninvoice.vn/\n'], ['0102182292', 'https://hddt.vnpay.vn/Invoice/Index/\n'], ['0106870211', 'https://tracuu.vietinvoice.vn/#/\n'], ['0104614692', 'https://hoadontvan.com/TraCuu\n'], ['0309612872', 'https://ehd.smartvas.vn/HDDT/\n'], ['0309478306', 'https://tracuu.xuathoadon.vn/\n'], ['0315298333', 'https://tctinvoice.com/\n'], ['0303609305', 'https://ihoadondientu.com/Tra-cuu\n'], ['0100727825', 'https://invoice.fast.com.vn/lookup/tra-cuu-hoa-don-dien-tu.aspx\n'], ['0315467091', 'http://www.acconine.vn ( không tồn tại website )\n'], ['0315638251', 'https://laphoadon.htinvoice.vn/TraCuu\n'], ['0105958921', 'https://tracuu.cloudinvoice.vn/\n'], ['0302431595', 'https://tracuu.hoadon30s.vn/en/tin-tuc/\n'], ['0103018807', 'https://vnisc.com.vn\n'], ['0106820789', 'https://tracuu.hoadondientuvn.info/#/tracuuhoadon/tracuu\n'], ['0310151055', 'https://www.SAFEinvoice.vn ( website không thể truy cập )\n'], ['0303430876', 'www.spc-technology.com ( website không thể truy cập ) \n'], ['0301452923', 'https://tracuu.lienson.vn/#/tracuuhoadon/tracuu\n'], ['0314185087', 'https://hoadon.onlinevina.com.vn/invoice\n'], ['0100687474', 'https://hoadondientu-ptp.vn/tra-cuu/\n'], ['0400462489', 'https://e-invoicetuanchau.com/Tra-cuu\n'], ['3500456910', 'https://hoadonminhthuvungtau.com/Tra-cuu\n'], ['0104908371', 'https://hoadondientu.acman.vn/tra-cuu/hoa-don.html\n'], ['0315191291', 'https://hoadonsovn.evat.vn/\n'], ['0313844107', 'http://voice.hoadondientu.net.vn\n'], ['0311622035', 'http://voice.hoadondientu.net.vn/tra-cuu\n'], ['0106361479', 'https://tracuu.ahoadon.com/\n'], ['0312270160', 'https://ameinvoice.vn/invoice-inquiry/\n'], ['0104493085', 'https://fts.com.vn/phan-mem-hoa-don-dien-tu/\n'], ['0101289966', 'https://tracuu.e-hoadon.cloud/\n'], ['0303211948', 'https://vlc.evat.vn/\n'], ['0101622374', 'https://tamvietgroup.vn/hoa-don-dien-tu/\n'], ['0310768095', 'http://hoadondientu.link/tracuu\n'], ['0312961577', 'http://tracuuhoadon.benthanhinvoice.vn/\n'], ['0313950909', 'https://koffi.vn\n'], ['0311928954', 'https://tracuuhoadon.vietinfo.tech/\n'], ['0103770970', 'https://www.bitware.vn/tracuuhoadon/\n'], ['0305142231', 'https://www.rosysoft.vn/tin-cong-nghe/erp-rosy-giai-phap-hoa-don-dien-tu\n'], ['3702037020', 'https://trandinhtung.evat.vn/\n'], ['0101925883', 'http://tracuu.cmcsoft.com/\n'], ['0316642395', 'https://phuongnam.evat.vn/\n'], ['0315194912', 'https://ttltax.com/dich-vu/hoa-don-dien-tu-237.html\n'], ['0315983667', 'http://hoadondientuvietnam.vn/HDDT/\n'], ['0310926922', 'https://invoice.ehcm.vn/\n'], ['0101010702', 'https://www.thanglongsoft.com/index.php\n'], ['0102720409', 'http://tigtax.vn\n'], ['0314058603', 'https://portal.vdsg-invoice.vn/\n'], ['0301448733', 'https://accnet.vn/hoa-don-dien-tu\n'], ['0313253288', 'https://app.autoinvoice.vn/tracuu.zul\n'], ['0309889835', 'https://unit.com.vn/'], ['0202029650', 'https://hdbk.pmbk.vn/tra-cuu-hoa-don'], ['0108971656', 'https://tracuu.myinvoice.vn/#/'], ['0312942260', 'https://ihoadondientu.net/Tracuu.aspx\n'], ['1201496252', 'https://webcashvietnam.com/vn/e_invoice.html\n'], ['0303549303', 'https://e-invoices.vn/\n'], ['0311914694', 'https://brightbrain.vn/?s=tra+c%E1%BB%A9u\n'], ['0312617990', 'https://demo-eportal.cloudteam.vn/n'], ['0109282176', 'https://tracuu.vininvoice.vn/\n'], ['0102723181', 'http://hoadonct.gov.vn/'], ['0106858609', 'https://tracuuhoadon.vetc.com.vn/\n'], ['0315151651', 'https://ei.pvssolution.com/#/\n'], ['0310151739', 'https://news.yoinvoice.vn\n'], ['0312575123', 'https://www.ecount.com/vn/ecount/product/accounting_e-invoice'], ['0107732197', 'https://tracuuhoadon.atis.com.vn/\n'], ['0101659906', 'https://tracuu.kaike.vn/#/\n'], ['0103019524', 'https://einvoice.aits.vn/\n'], ['0316114998', 'https://bizzi.vn/\n'], ['0316636497', 'http://beetek.vn/\n'],]
        
        try:
            new_directory_name = self.user
            folder_son = os.path.join(self.path_ketqua, new_directory_name)
            if os.path.exists(self.path_ketqua):
                if os.path.exists(folder_son):
                    pass
                else:
                    os.makedirs(folder_son)
                    print("Thư mục con đã được tạo thành công.")
            else:
                return {
                    'success': False,
                    'message': 'Thư mục không tồn tại !'
                }

            if self.xml_:
                folder_path_xml = os.path.join(folder_son ,f'XML {self.begin_.replace("/","-")}_{self.end_.replace("/","-")}')
                if not os.path.exists(folder_path_xml):
                    os.makedirs(folder_path_xml)

            if self.html_:
                folder_path_html = os.path.join(folder_son ,f'HTML {self.begin_.replace("/","-")}_{self.end_.replace("/","-")}')
                if not os.path.exists(folder_path_html):
                    os.makedirs(folder_path_html)

            excel_thongke_a = '.\\__pycache__\\cache_\\template\\Thống kê chi tiết.xlsx'
            full_path_a = os.path.abspath(excel_thongke_a)
            full_path_b = folder_son
            shutil.copy(full_path_a, full_path_b)
            name_f = os.path.join(full_path_b,'Thống kê chi tiết.xlsx')
            day__ = self.range_DAY_const.replace("Thời gian : ","").replace("/","-").replace("=>","_")
            us = self.user.replace("\r","")
            br = 1
            if type == 1:
                br = 2
                mst = 1
                type_mb = self.user
                type_hoadon = 'sold'
                name_f1= os.path.join(folder_son, (us +'_HDCTBanra' + day__  + '.xlsx'))
                q = 0
                try:
                    os.rename(name_f, name_f1)
                except:
                    return {
                        'success': False,
                        'message': 'Đã tồn tại thống kê cho thời gian này !'
                    }
            d = ""
            e = ""
            type_list = {""}
            if type == 2:
                mst = 2
                type_mb = "nbmst"
                type_list = {'5','6','8'}
                d = f';ttxly=='
                type_hoadon = 'purchase'
                name_f1= os.path.join(folder_son, (us +'_HDCTMuaVao' + day__  + '.xlsx'))
                q = 0
                try:
                    os.rename(name_f, name_f1)
                except:
                    return {
                        'success': False,
                        'message': 'Đã tồn tại thống kê cho thời gian này !'
                    }

            self.len_a = len(self.arr_ed)
            self.n_pro = int(100/self.len_a)
            self.a = 0
            datas_first = ""
            n_collect = 100/len(self.arr_ed)
            for self.i in range(len(self.arr_ed)):
                print("RUN")
                self.a += n_collect
                begin_day = self.arr_ed[self.i][0]
                end_day = self.arr_ed[self.i][1]
                print(begin_day)
                print(end_day)
                spec = ""
                for e in type_list:
                    for i in range(br):
                        if e == "8":
                            spec = "sco-"
                        else:
                            spec = ""
                        if i == 1:
                            spec = "sco-"
                            print(2)    
                        while True:
                            try:
                                res = self.ssl_supressed_session().get(f'https://hoadondientu.gdt.gov.vn:30000/{spec}query/invoices/{type_hoadon}?sort=tdlap:desc,khmshdon:asc,shdon:desc&size=50&search=tdlap=ge={begin_day}T00:00:00;tdlap=le={end_day}T23:59:59{d}{e}',headers=self.headers,verify=False,timeout=1)
                                if res.status_code == 200:
                                    break
                            except:
                                print("error")
                        data = res.json()
                        if datas_first == "":
                            print("first")
                            datas_first = data
                        else:
                            print("add ")
                            datas_first["datas"].extend(data["datas"])
                        while True:
                            if data["state"] != None and data != "":
                                print("chèn")
                                print(data['state'])
                                while True:
                                    try:
                                        res = self.ssl_supressed_session().get(f'https://hoadondientu.gdt.gov.vn:30000/{spec}query/invoices/{type_hoadon}?sort=tdlap:desc,khmshdon:asc,shdon:desc&size=50&state={data["state"]}&search=tdlap=ge={begin_day}T00:00:00;tdlap=le={end_day}T23:59:59',headers=self.headers,verify=False,timeout=1)
                                        if res.status_code == 200:
                                            break
                                        else:
                                            pass
                                    except:
                                        print("error")
                                data = res.json()
                                if "state" in data:
                                    datas_first["datas"].extend(data["datas"])
                            else:
                                break
            datas_first["datas"] = self.remove_duplicate_elements(datas_first["datas"])
            count = len(datas_first["datas"])
            wb = openpyxl.load_workbook(name_f1)
            sheet = wb.active
            last_row = sheet.max_row
            start_index = 1
            hdon = {1: "Hóa đơn mới", 2: "Hóa đơn thay thế", 3: "Hóa đơn điều chỉnh", 4: "Hóa đơn đã bị thay thế", 5: "Hóa đơn đã bị điều chỉnh", 6: "Hóa đơn đã bị hủy"}
            ttxly = {0: "Tổng cục Thuế đã nhận", 1: "Đang tiến hành kiểm tra điều kiện cấp mã", 2: "CQT từ chối hóa đơn theo từng lần phát sinh", 3: "Hóa đơn đủ điều kiện cấp mã", 4: "Hóa đơn không đủ điều kiện cấp mã", 5: "Đã cấp mã hóa đơn", 6: "Tổng cục thuế đã nhận không mã", 7: "Đã kiểm tra định kỳ HĐĐT không có mã", 8: "Tổng cục thuế đã nhận hóa đơn có mã khởi tạo từ máy tính tiền"}
            self.n_pro = int(100/self.len_a)
            self.a = 0
            n_begin_load = 0
            len_hd = len(datas_first["datas"])
            try:    
                n_collect = 100/len_hd
            except:
                pass
            start_index = 0
            dem = 2
            spec = ""
            temp_ = []
            for data in datas_first["datas"]:
                if data["ttxly"] == 8:
                    spec = "sco-"
                else:
                    spec = ""
                start_index+=1
                nbmst = data["nbmst"]
                nmmst = data["nmmst"]
                khhdon = data["khhdon"]
                shd = data["shdon"]
                khmshdon = data["khmshdon"]
                nbd = data["ntao"].split("T")[0]
                if self.checkspc == 'mv':
                    mst = data["nbmst"]
                elif self.checkspc == 'br':
                    mst = data["nmmst"]
                time.sleep(0.2)
                while True:    
                    try:                     
                        res1 = self.ssl_supressed_session().get(f'https://hoadondientu.gdt.gov.vn:30000/{spec}query/invoices/detail?nbmst={nbmst}&khhdon={khhdon}&shdon={shd}&khmshdon={khmshdon}',headers=self.headers,verify=False,timeout =1)
                        if res1.status_code == 200:
                            break
                    except:
                        print("ERROR")
                data_ct = res1.json()
                zip_data = res1.content
                if self.xml_:
                    xml = os.path.join(folder_path_xml , f"{khmshdon}_{khhdon}_{shd}_{nbd}_{nbmst}.xml")  
                    if self.banra:
                        xml = os.path.join(folder_path_xml , f"{khmshdon}_{khhdon}_{shd}_{nbd}_{nmmst}.xml")                                
                    if not os.path.exists(folder_path_xml):
                        os.makedirs(folder_path_xml)
                    try:
                        with zipfile.ZipFile(zip_data, "r") as zip_file:
                            for filename in zip_file.namelist():
                                if filename == "invoice.xml":
                                    file_content = zip_file.read('invoice.xml').decode('utf-8')
                                    with open(xml, "w",encoding="utf-8-sig") as output_file:
                                        output_file.write(file_content)
                    except:
                        print(zip_data)
                        pass
                if self.html_:                          
                    html = os.path.join(folder_path_html , f"{khmshdon}_{khhdon}_{shd}_{nbd}_{nbmst}.html")
                    if self.banra:
                        html = os.path.join(folder_path_html , f"{khmshdon}_{khhdon}_{shd}_{nbd}_{nmmst}.html")
                    if not os.path.exists(folder_path_html):
                        os.makedirs(folder_path_html)
                    try:
                        with zipfile.ZipFile(zip_data, "r") as zip_file:
                            for filename in zip_file.namelist():
                                if filename == "invoice.html":
                                    file_content = zip_file.read('invoice.html').decode('utf-8')
                                    with open(html, "w",encoding="utf-8-sig") as output_file:
                                        output_file.write(file_content)
                    except:
                        print(zip_data)
                        pass
                headers = ["khmshdon","khhdon","shdon","ntao","nky","mhdon","nky","dvtte","tgia","nbten", "nbmst","nbdchi","nmten","nmmst","nmdchi","m_VT","ten","dvtinh","sluong","dgia","stckhau","tsuat","thtien","tthue","ttcktmai","tgtphi","tgtttbso","tthai","ttxly","url","mk","ghichu","thtttoan","tchat","dgiai"]
                s = 0
                n = 0
                url = ""
                mk = ""
                if mk == "":
                    try:
                        for i in data_ct["ttkhac"]:
                            if i["ttruong"] == "Mã số bí mật" or i["ttruong"] == "KeySearch" or i["ttruong"] == "Mã TC" or i["ttruong"] == "TransactionID" or i["ttruong"] == "Fkey" or i["ttruong"] == "MNHDon" or i["ttruong"] == "QuanLy_SoBaoMat" or i["ttruong"] == "Mã bảo mật" or i["ttruong"] == "Số bảo mật" or i["ttruong"] == "Mã tra cứu hóa đơn" or i["ttruong"] == "chungTuLienQuan"  or i["ttruong"] == "InvoiceId" or i["ttruong"] == "MaTraCuu" or i["ttruong"] == "MTCuu" or i["ttruong"] == "SearchInvoice" or i["ttruong"] == "Mã tra cứu" :
                                mk = i["dlieu"]
                        if mk == "":
                            for i in data_ct["cttkhac"]:
                                if i["ttruong"] == "Mã số bí mật" or i["ttruong"] == "KeySearch" or i["ttruong"] == "Mã TC" or i["ttruong"] == "TransactionID" or i["ttruong"] == "Fkey" or i["ttruong"] == "MNHDon" or i["ttruong"] == "QuanLy_SoBaoMat" or i["ttruong"] == "Mã bảo mật" or i["ttruong"] == "Số bảo mật" or i["ttruong"] == "Mã tra cứu hóa đơn"  or i["ttruong"] == "chungTuLienQuan" or i["ttruong"] == "InvoiceId" or i["ttruong"] == "MaTraCuu" or i["ttruong"] == "MTCuu" or i["ttruong"] == "SearchInvoice" or i["ttruong"] == "Mã tra cứu" :
                                    mk = i["dlieu"]
                        if mk == "":
                            for i in data_ct["ttttkhac"]:
                                if i["ttruong"] == "Mã số bí mật" or i["ttruong"] == "KeySearch" or i["ttruong"] == "Mã TC" or i["ttruong"] == "TransactionID" or i["ttruong"] == "Fkey" or i["ttruong"] == "MNHDon" or i["ttruong"] == "QuanLy_SoBaoMat" or i["ttruong"] == "Mã bảo mật" or i["ttruong"] == "Số bảo mật" or i["ttruong"] == "chungTuLienQuan" or i["ttruong"] == "Mã tra cứu hóa đơn" or i["ttruong"] == "InvoiceId" or i["ttruong"] == "MaTraCuu" or i["ttruong"] == "MTCuu" or i["ttruong"] == "SearchInvoice" or i["ttruong"] == "Mã tra cứu" :
                                    mk = i["dlieu"]
                        try:
                            if mk == "":
                                for i in data_ct["TTKhac"]:
                                    if i["ttruong"] == "Mã số bí mật" or i["ttruong"] == "KeySearch" or i["ttruong"] == "Mã TC" or i["ttruong"] == "TransactionID" or i["ttruong"] == "Fkey" or i["ttruong"] == "MNHDon" or i["ttruong"] == "QuanLy_SoBaoMat" or i["ttruong"] == "Mã bảo mật" or i["ttruong"] == "Số bảo mật" or i["ttruong"] == "chungTuLienQuan" or i["ttruong"] == "Mã tra cứu hóa đơn" or i["ttruong"] == "InvoiceId" or i["ttruong"] == "MaTraCuu" or i["ttruong"] == "MTCuu" or i["ttruong"] == "SearchInvoice":
                                        mk = i["dlieu"]
                        except:
                            pass
                    except:
                        pass
                if mk == "":
                    mk = "Không tìm thấy mã tra cứu trên file XML, vui lòng liên hệ người bán để được cung cấp file PDF gốc."
                if url == "https://tracuu.vietinvoice.vn/#/" or url == "https://tracuu.vininvoice.vn":
                    mk = data_ct["mhdon"]
                if url == "https://lottemart-nsg-tt78.vnpt-invoice.com.vn/Portal/Index":
                    mk = "Tìm kiếm theo thông tin hóa đơn"
                if url == "EASY":
                    try:
                        fkey = data['mhdon']
                    except Exception as e:
                        print(e)
                        time.sleep(10000)
                    url = f"https://{nbmst}hd.easyinvoice.com.vn/Search/?strFkey={fkey}"
                d = 0
                z = 0
                p = 0
                ttttoan = ""
                tong_tt = 0
                tong_ttct = 0 
                asss = 0 
                while True:
                    t = 0
                    temp = ""
                    tong_thue = 0 
                    try:    
                        len_ct = len(data_ct["hdhhdvu"])
                    except:
                        break
                    sp_n = 0
                    for sttt,sp in enumerate(data_ct["hdhhdvu"]):  
                        sp_n +=1
                        if z == 0:
                            ttttoan = "tgtttbso"
                            z +=1
                        else:
                            ttttoan = ""
                            z+=1  
                        row_index = last_row + 1
# --- BẮT ĐẦU DÁN MÃ MỚI VÀO ĐÂY ---
                        values = []  # Khởi tạo lại list giá trị cho mỗi dòng hàng hóa
                        # Biến temp để kiểm tra và chỉ ghi một số giá trị tổng hợp ở dòng đầu tiên (như func1)
                        temp_shdon_check = data_ct.get("shdon", "")
# --- BẮT ĐẦU THAY THẾ/DÁN ĐOẠN MÃ NÀY ---

                        # Chuẩn bị trước các giá trị ngày tháng từ data và data_ct
                        ntao_val = data.get("tdlap", "")
                        nky_seller_val = data.get("nky", "")
                        nky_cqt_val = data_ct.get("nky", "") # Ngày CQT ký từ chi tiết

                        # Hàm nội bộ để định dạng ngày (tránh lặp code)
                        def format_date(date_str):
                            if not date_str: return ""
                            try:
                                date_part = date_str.split('T')[0]
                                parts = date_part.split('-')
                                if len(parts) == 3:
                                    return f"{parts[2]}/{parts[1]}/{parts[0]}"
                                return date_part # Trả về nếu không đúng format YYYY-MM-DD
                            except:
                                return date_str # Trả về gốc nếu có lỗi

                        ntao_formatted = format_date(ntao_val)
                        nky_seller_formatted = format_date(nky_seller_val)
                        if not nky_seller_formatted:  # Fallback nếu ngày ký người bán trống
                            nky_seller_formatted = ntao_formatted
                        nky_cqt_formatted = format_date(nky_cqt_val)

                        # Kiểm tra xem hóa đơn có chi tiết hàng hóa không
                        has_items = "hdhhdvu" in data_ct and data_ct["hdhhdvu"]

                        if has_items:
                            # --- XỬ LÝ HÓA ĐƠN CÓ CHI TIẾT HÀNG HÓA ---
                            z = 0  # Biến đếm dòng hàng hóa, bắt đầu từ 0 cho mỗi hóa đơn
                            tong_thue_invoice = 0.0 # Biến lưu tổng tiền thuế đã ghi ra cho các dòng trước đó của hóa đơn này
                            len_ct = len(data_ct["hdhhdvu"]) # Số lượng dòng chi tiết

                            for sttt, sp in enumerate(data_ct["hdhhdvu"]): # Lặp qua từng dòng hàng hóa
                                row_index = last_row + 1
                                values = []

                                for n, header in enumerate(headers): # Lặp qua từng cột trong header
                                    value = "" # Giá trị mặc định

                                    # Lấy giá trị từ data_ct (thông tin chung của hóa đơn)
                                    if header in ["khmshdon", "khhdon", "shdon", "mhdon", "dvtte", "tgia", "nbten", "nbmst", "nbdchi", "nmten", "nmmst", "nmdchi", "ttcktmai", "tgtphi", "tgtttbso", "tthai", "ttxly", "thtttoan"]:
                                        if header in ["tgtttbso", "tgtphi", "ttcktmai"]: # Các trường tổng chỉ ghi ở dòng đầu (z=0)
                                            if z == 0: value = data_ct.get(header, "")
                                            else: value = ""
                                        elif header == 'nbten' and (data_ct.get("nbten", "") == 'null' or data_ct.get("nbten", "") == None): value = data_ct.get("nmtnban", "")
                                        elif header == 'nmten' and (data_ct.get("nmten", "") == 'null' or data_ct.get("nmten", "") == None): value = data_ct.get("nmtnmua", "")
                                        else: value = data_ct.get(header, "")

                                        # Chuyển đổi mã trạng thái
                                        if header == "tthai":
                                            if value in hdon: value = hdon[value]
                                        if header == "ttxly":
                                            if value in ttxly: value = ttxly[value]

                                    # Gán giá trị ngày tháng đã định dạng
                                    elif header == "ntao": value = ntao_formatted
                                    elif header == "nky" and n == 4: value = nky_seller_formatted # Index 4 là ngày người bán ký
                                    elif header == "nky" and n == 6: value = nky_cqt_formatted # Index 6 là ngày CQT ký

                                    # URL, Mã tra cứu chỉ ghi ở dòng đầu (z=0)
                                    elif header == "url":
                                        if z == 0: value = url
                                        else: value = ""
                                    elif header == "mk":
                                        if z == 0: value = mk
                                        else: value = ""

                                    # Lấy giá trị từ sp (chi tiết từng dòng hàng hóa)
                                    elif header in ["m_VT","ten","dvtinh","sluong","dgia","stckhau","tsuat","thtien","tthue","tchat"]:
                                        value = sp.get(header, "")
                                        if header == "tsuat": # Xử lý hiển thị thuế suất
                                            ltsuat = sp.get("ltsuat", "")
                                            if ltsuat == "KHAC" : value = "KHAC"
                                            elif ltsuat == "KKKNT": value = "KKKNT"
                                            elif isinstance(value, (int, float)) and value == 0.0 and ltsuat == "KCT": value = "KCT"
                                            elif isinstance(value, (int, float)) and value == 0.0: value = "0%"
                                            elif isinstance(value, (int, float)): value = f"{value * 100}%" # Hiển thị dạng %

                                        if header == "tchat": # Xử lý hiển thị tính chất
                                            if value == 1: value = "Hàng hóa, dịch vụ"
                                            elif value == 2: value = "Khuyến mại"
                                            elif value == 3: value = "Chiết khấu"
                                            elif value == 4: value = "Ghi chú, diễn giải"

                                    # Ghi chú và Diễn giải chỉ ghi ở dòng đầu (z=0)
                                    elif header == "ghichu":
                                        if z == 0:
                                            try: # So sánh ngày lập và ngày ký người bán
                                                d1_g = ntao_formatted.split("/")
                                                d2_g = nky_seller_formatted.split("/")
                                                if len(d1_g)==3 and len(d2_g)==3 and (d1_g != d2_g): value = "  " # Đánh dấu nếu khác nhau
                                                else: value = " "
                                            except: value = " "
                                        else: value = " "
                                    elif header == "dgiai":
                                        if z == 0:
                                            # Tạo diễn giải cho hóa đơn điều chỉnh/thay thế
                                            khmshd = data_ct.get("khmshdgoc", "")
                                            khhd = data_ct.get("khhdgoc", "")
                                            shd_ = data_ct.get("shdgoc", "")
                                            tdlhdgoc = data_ct.get("tdlhdgoc", None)
                                            if tdlhdgoc and data_ct.get("tthai") in [2, 3]: # Nếu là HĐ thay thế/điều chỉnh
                                                nlap = format_date(tdlhdgoc)
                                                value = f"Điều chỉnh/Thay thế cho HĐ {khmshd}{khhd}, số {shd_}, ngày {nlap}"
                                            else: # Nếu không phải, thử lấy ghi chú chung
                                                value = " "
                                                try:
                                                    for i in data_ct.get("ttkhac", []):
                                                        if i.get("ttruong") == "Ghi chú hóa đơn":
                                                            dlieu = i.get("dlieu", "")
                                                            if dlieu and len(dlieu) > 1: # Lấy ghi chú nếu có và đủ dài
                                                                 value = dlieu
                                                                 break
                                                except: pass
                                        else: value = " "
                                    else: # Header không xác định hoặc là cột phụ (Số lô, Hạn dùng)
                                         # Sẽ được xử lý sau vòng lặp này
                                         pass # Để trống value ở bước này

                                    values.append(value) # Thêm giá trị vào list của hàng hiện tại

                                # --- Xử lý sau khi đã có đủ giá trị cơ bản cho hàng ---
                                # Tính toán/Điều chỉnh tiền thuế (Cần logic phức tạp giống file 1 để khớp tổng)
                                try:
                                    tsuat_idx = headers.index("tsuat")
                                    tthue_idx = headers.index("tthue")
                                    thtien_idx = headers.index("thtien")
                                    tgtthue_total = float(data_ct.get("tgtthue", 0) or 0) # Tổng thuế thực tế của hóa đơn

                                    current_thtien = float(values[thtien_idx] or 0)
                                    current_tsuat_str = str(values[tsuat_idx])

                                    calculated_tax = 0.0
                                    if "%" in current_tsuat_str:
                                        calculated_tax = current_thtien * (float(current_tsuat_str.replace('%','')) / 100.0)
                                    elif current_tsuat_str == "KHAC":
                                        calculated_tax = current_thtien * (5.263 / 100.0) # Giả định KHAC là 5.263%

                                    # Làm tròn tiền thuế (ví dụ làm tròn đến 2 chữ số thập phân)
                                    calculated_tax = round(calculated_tax, 2)

                                    # Nếu là dòng cuối cùng, điều chỉnh để khớp tổng thuế
                                    if sttt == len_ct - 1:
                                        calculated_tax = tgtthue_total - tong_thue_invoice
                                    else:
                                        tong_thue_invoice += calculated_tax

                                    values[tthue_idx] = calculated_tax # Gán tiền thuế đã tính/điều chỉnh
                                except Exception as tax_err:
                                    print(f"Lỗi tính thuế HĐ {data.get('shdon')}: {tax_err}, giá trị gốc: {sp.get('tthue', '')}")
                                    try: values[tthue_idx] = float(sp.get('tthue', 0) or 0) # Thử dùng giá trị gốc nếu lỗi
                                    except: values[tthue_idx] = 0


                                # Điều chỉnh thành tiền âm nếu là chiết khấu
                                try:
                                    tchat_idx = headers.index("tchat")
                                    thtien_idx = headers.index("thtien")
                                    if values[tchat_idx] == "Chiết khấu":
                                        values[thtien_idx] = abs(float(values[thtien_idx] or 0)) * -1
                                except Exception as ck_err: print(f"Lỗi CK HĐ {data.get('shdon')}: {ck_err}")

                                # Định dạng lại số hóa đơn
                                try:
                                     shd_idx = headers.index("shdon")
                                     values[shd_idx] = " " + str(values[shd_idx])
                                except IndexError: pass

                                # Thêm Số lô, Hạn dùng (nếu có) vào cuối list values
                                lot_val, exp_val = "", ""
                                try:
                                     for i in sp.get("ttkhac", []):
                                         if i.get("ttruong") == "Lot": lot_val = i.get("dlieu", "")
                                         if i.get("ttruong") == "ExpireDate": exp_val = i.get("dlieu", "")
                                except: pass
                                # Kiểm tra xem header có cột này không trước khi append
                                if "Số lô" in headers: values.append(lot_val)
                                if "Hạn dùng" in headers: values.append(exp_val)


                                # Ghi dữ liệu hàng này vào sheet
                                for column_index, cell_value in enumerate(values, start=1):
                                     # Giới hạn số cột ghi ra bằng số lượng header tối đa có thể có
                                     # (số header chuẩn + số cột phụ như Số lô, Hạn dùng)
                                     max_cols = len(headers) # Điều chỉnh nếu có cột phụ
                                     if "Số lô" in headers: max_cols +=1
                                     if "Hạn dùng" in headers: max_cols +=1

                                     if column_index <= max_cols:
                                        sheet.cell(row=row_index, column=column_index, value=cell_value)
                                last_row += 1
                                z += 1 # Tăng biến đếm dòng hàng hóa đã xử lý

                        else:
                            # --- XỬ LÝ HÓA ĐƠN KHÔNG CÓ CHI TIẾT HÀNG HÓA ---
                            print(f"HĐ {data.get('khmshdon')}_{data.get('khhdon')}_{data.get('shdon')} không có hdhhdvu. Ghi dòng tổng hợp.")
                            row_index = last_row + 1
                            values = []

                            for n, header in enumerate(headers): # Lặp qua các cột
                                value = ""
                                # Lấy thông tin chung từ data_ct
                                if header in ["khmshdon", "khhdon", "shdon", "mhdon", "dvtte", "tgia", "nbten", "nbmst", "nbdchi", "nmten", "nmmst", "nmdchi", "ttcktmai", "tgtphi", "tgtttbso", "tthai", "ttxly", "thtttoan"]:
                                    if header == 'nbten' and (data_ct.get("nbten", "") == 'null' or data_ct.get("nbten", "") == None): value = data_ct.get("nmtnban", "")
                                    elif header == 'nmten' and (data_ct.get("nmten", "") == 'null' or data_ct.get("nmten", "") == None): value = data_ct.get("nmtnmua", "")
                                    else: value = data_ct.get(header, "")
                                    if header == "tthai":
                                        if value in hdon: value = hdon[value]
                                    if header == "ttxly":
                                        if value in ttxly: value = ttxly[value]
                                # Lấy ngày tháng đã định dạng
                                elif header == "ntao": value = ntao_formatted
                                elif header == "nky" and n == 4: value = nky_seller_formatted
                                elif header == "nky" and n == 6: value = nky_cqt_formatted
                                # Lấy URL, Mã tra cứu
                                elif header == "url": value = url
                                elif header == "mk": value = mk
                                # Các cột chi tiết hàng hóa để trống
                                elif header in ["m_VT","ten","dvtinh","sluong","dgia","stckhau","tsuat","thtien","tthue","tchat"]:
                                    value = ""
                                # Ghi chú và diễn giải
                                elif header == "ghichu":
                                     try: # So sánh ngày
                                         d1_g = ntao_formatted.split("/")
                                         d2_g = nky_seller_formatted.split("/")
                                         if len(d1_g)==3 and len(d2_g)==3 and (d1_g != d2_g): value = "  "
                                         else: value = " "
                                     except: value = " "
                                elif header == "dgiai":
                                    # Logic diễn giải tương tự như trên
                                    khmshd = data_ct.get("khmshdgoc", "")
                                    khhd = data_ct.get("khhdgoc", "")
                                    # ... (lấy các thông tin khác) ...
                                    if data_ct.get("tdlhdgoc", None) and data_ct.get("tthai") in [2, 3]:
                                        value = f"Điều chỉnh/Thay thế cho HĐ ..."
                                    else: # Lấy ghi chú chung nếu có
                                        value = " "
                                        try:
                                            for i in data_ct.get("ttkhac", []):
                                                 if i.get("ttruong") == "Ghi chú hóa đơn":
                                                      # ... (lấy ghi chú) ...
                                                      break
                                        except: pass
                                else: # Header không xác định
                                    value = data_ct.get(header, "") # Thử lấy từ data_ct

                                values.append(value)

                            # Thêm cột trống cho Số lô, Hạn dùng nếu có trong header
                            if "Số lô" in headers: values.append("")
                            if "Hạn dùng" in headers: values.append("")

                            # Định dạng lại số hóa đơn
                            try:
                                 shd_idx = headers.index("shdon")
                                 values[shd_idx] = " " + str(values[shd_idx])
                            except IndexError: pass

                            # Ghi dòng tổng hợp này vào sheet
                            for column_index, cell_value in enumerate(values, start=1):
                                 max_cols = len(headers)
                                 if "Số lô" in headers: max_cols +=1
                                 if "Hạn dùng" in headers: max_cols +=1
                                 if column_index <= max_cols:
                                     sheet.cell(row=row_index, column=column_index, value=cell_value)
                            last_row += 1 # Tăng last_row sau khi ghi dòng tổng hợp

                        # --- KẾT THÚC PHẦN DÁN MÃ MỚI ---
                
                    break

            border = Border(left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"))
            columns = ['H','I','J','K','L']
            for column in columns:
                for cell in sheet[column]:
                    cell.number_format = '#,##0'
            alignment = Alignment(wrap_text=True)
            for row in sheet.iter_rows(min_row=7, max_row=last_row, min_col=1, max_col=16):
                for cell in row:
                    cell.alignment = alignment
            font = Font(size=12)
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = font
            columns = {"E"}
            for column in columns:
                for cell in sheet[column]:
                    try:
                        new_value = cell.value
                        new_value = new_value.split("-")
                        print(new_value)
                        new_value = new_value[2] + "/" + new_value[1] + "/" + new_value[0]
                        cell.value = new_value
                    except:
                        pass
            wb.save(name_f1)
              # Convert workbook to JSON and return
            json_data = self.workbook_to_json(wb)
            print(json_data)
            
            return {
                'success': True,
                'message': f'Đã xử lý thành công {start_index}/{len_hd} hóa đơn',
                'file_path': name_f1,                'json_data': json_data
            }
        except Exception as e:
            return {
                'success': False,
                'message': str(e)
            }

    def save_to_mongodb(self, data, invoice_type):
        """Save invoice data to MongoDB without duplicate checking
        invoice_type: 'sale' or 'purchase'"""
        try:
            # Select the appropriate collection based on invoice type
            collection = self.sales_collection if invoice_type == 'sale' else self.purchase_collection
            
            print(f"Starting to save {len(data)} invoices to MongoDB for {invoice_type}")
            
            # Insert all invoices without checking for duplicates
            inserted_count = 0
            for invoice in data:
                # Insert new invoice
                result = collection.insert_one(invoice)
                inserted_count += 1
                print(f"✓ Inserted {invoice_type} invoice: {invoice.get('Số hóa đơn', 'N/A')} - ID: {result.inserted_id}")
            
            print(f"MongoDB save completed: {inserted_count} invoices inserted")
            
            return {
                'success': True,
                'message': f'{invoice_type.capitalize()} data saved to MongoDB successfully - {inserted_count} invoices inserted'
            }
        except Exception as e:
            print(f"Error in save_to_mongodb: {str(e)}")
            return {
                'success': False,
                'message': f'Error saving {invoice_type} data to MongoDB: {str(e)}'
            }

    def workbook_to_json(self, wb):
        """Convert workbook data to JSON and return it"""
        sheet = wb.active
        headers = []
        data = []
        
        # Get headers from first row
        for cell in sheet[1]:
            headers.append(cell.value)
            
        # Get data from remaining rows
        for row in sheet.iter_rows(min_row=2):
            row_data = {}
            for idx, cell in enumerate(row):
                if idx < len(headers):
                    row_data[headers[idx]] = cell.value
            # Add username to each row
            row_data['username'] = self.user
            data.append(row_data)
            
        # Save to MongoDB - determine invoice type based on self.banra or self.muavao
        invoice_type = 'sale' if self.banra else 'purchase'
        self.save_to_mongodb(data, invoice_type)
            
        # Return Python object instead of JSON string
        return data

class CustomHttpAdapter(adapters.HTTPAdapter):
    def __init__(self, ssl_context=None, **kwargs):
        self.ssl_context = ssl_context
        super().__init__(**kwargs)
    def init_poolmanager(self, connections, maxsize, block=False):
        self.poolmanager = poolmanager.PoolManager(
            num_pools=connections, maxsize=maxsize,
            block=block, ssl_context=self.ssl_context) 
